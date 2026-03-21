#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "delivery" / "09-project-acceptance-checklist.xlsx"


def build_styles():
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    done_fill = PatternFill("solid", fgColor="DCFCE7")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header_fill": header_fill,
        "done_fill": done_fill,
        "border": border,
        "header_font": Font(color="FFFFFF", bold=True, size=11),
        "cell_font": Font(color="111827", size=10),
        "title_font": Font(color="111827", bold=True, size=14),
    }


def write_overview(ws, styles):
    ws.title = "项目总览"
    ws["A1"] = "Demand Intelligence Platform｜项目交付总览"
    ws["A1"].font = styles["title_font"]
    ws.merge_cells("A1:E1")

    rows = [
        ("项目名称", "Demand Intelligence Platform"),
        ("项目类型", "需求情报决策系统 / 混合部署产品"),
        ("交付状态", "第一阶段完整闭环已完成"),
        ("GitHub", "https://github.com/Perrilee711/raddit"),
        ("团队工作台", "http://43.162.90.26/"),
        ("API 健康检查", "http://43.162.90.26/api/health"),
        ("Vercel 展示版", "https://skill-deploy-jr9bh4v87v.vercel.app"),
    ]
    for idx, (key, value) in enumerate(rows, start=3):
        ws[f"A{idx}"] = key
        ws[f"B{idx}"] = value
        ws[f"A{idx}"].font = Font(bold=True)

    ws["A12"] = "管理层结论"
    ws["A12"].font = Font(color="111827", bold=True, size=11)
    ws["A13"] = "项目已形成业务-产品-设计-技术-测试-上线-运维的完整闭环，可作为后续项目复用样板。"
    ws.merge_cells("A13:E14")
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 24 if col == "A" else 30


def write_checklist(ws, styles):
    headers = ["编号", "分类", "验收项", "预期结果", "当前状态", "验证方式", "优先级", "备注"]
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=value)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ("F-01", "功能", "工作台可访问", "团队工作台首页可正常打开", "通过", "打开 http://43.162.90.26/", "P0", ""),
        ("F-02", "功能", "健康检查可访问", "/api/health 返回正常 JSON", "通过", "访问 http://43.162.90.26/api/health", "P0", ""),
        ("F-03", "功能", "Study 可创建", "Study Setup 可创建并首跑", "通过", "前端创建 Study", "P0", ""),
        ("F-04", "功能", "任务可入队", "browser/hot_threads/adaptive 均可入队", "通过", "Operations 查看", "P0", ""),
        ("F-05", "功能", "Mac Worker 可接单", "公网 API 发出的 mac_worker 任务被认领", "通过", "查看 jobs 与 worker 状态", "P0", ""),
        ("F-06", "功能", "结果可回写", "discover/harvest 结果写入 payload 并刷新 Dashboard", "通过", "刷新 Dashboard", "P0", ""),
        ("F-07", "功能", "停止当前爬取", "running 任务可进入 canceling/canceled", "通过", "点击 stop 按钮", "P0", ""),
        ("F-08", "功能", "失败任务可重试", "失败任务可重新排队运行", "通过", "Operations 重试", "P1", ""),
        ("O-01", "运维", "主服务状态可感知", "状态脚本输出清晰摘要", "通过", "执行 status_launch_agent.sh", "P1", ""),
        ("O-02", "运维", "Worker 状态可感知", "页面显示在线/离线/stale", "通过", "查看 Dashboard 状态卡", "P1", ""),
        ("O-03", "运维", "异常与告警提示", "失败任务与异常能在 Dashboard 呈现", "通过", "查看 Dashboard 告警区", "P1", ""),
        ("R-01", "发布", "GitHub 仓库可访问", "最新代码已推送至 main", "通过", "查看 GitHub 仓库", "P1", ""),
        ("R-02", "发布", "Vercel 展示版可访问", "展示壳可正常打开", "通过", "访问 Vercel 链接", "P2", "展示壳非完整操作入口"),
        ("B-01", "边界", "浏览器采集依赖 Mac", "Mac 睡眠/Chrome 权限异常会影响采集", "已确认", "运行手册说明", "P1", ""),
        ("B-02", "边界", "公网工作台当前为 HTTP", "HTTPS 与正式 API 域名待下一阶段补齐", "已确认", "发布文档说明", "P2", ""),
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.font = styles["cell_font"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 5 and value in {"通过", "已确认"}:
                cell.fill = styles["done_fill"]

    widths = [12, 10, 26, 36, 12, 24, 10, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"


def write_release(ws, styles):
    ws.title = "上线信息"
    headers = ["类型", "地址 / 入口", "说明"]
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=value)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("GitHub", "https://github.com/Perrilee711/raddit", "代码仓库"),
        ("团队工作台", "http://43.162.90.26/", "团队真实操作入口"),
        ("API 健康检查", "http://43.162.90.26/api/health", "云上 API 健康检查"),
        ("Vercel 展示版", "https://skill-deploy-jr9bh4v87v.vercel.app", "对外展示壳"),
        ("主服务状态脚本", "scripts/status_launch_agent.sh", "本机主服务运行摘要"),
        ("Worker 状态脚本", "scripts/status_mac_worker_launch_agent.sh", "Mac Worker 状态检查"),
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles["border"]
            cell.font = styles["cell_font"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, width in zip(["A", "B", "C"], [18, 48, 28]):
        ws.column_dimensions[col].width = width


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    styles = build_styles()
    wb = Workbook()
    overview = wb.active
    write_overview(overview, styles)
    write_checklist(wb.create_sheet("验收清单"), styles)
    write_release(wb.create_sheet("上线信息"), styles)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
